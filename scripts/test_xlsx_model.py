#!/usr/bin/env python3
"""Round-trip tests for scripts/xlsx_model.py."""

from __future__ import annotations

import shutil
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

sys.path.insert(0, str(Path(__file__).resolve().parent))
import xlsx_model as xm

BLUE_FILL = PatternFill(start_color="DEEAF6", end_color="DEEAF6", fill_type="solid")
BLUE_FONT = Font(color="0000FF")


def _paint_input(cell, value) -> None:
    cell.value = value
    cell.font = BLUE_FONT
    cell.fill = BLUE_FILL


def write_mini_workbook(path: Path) -> None:
    wb = Workbook()
    cover = wb.active
    cover.title = "00_Cover"
    cover["A1"] = "Price"
    cover["B1"] = "=Inputs!B1"
    cover["A2"] = "Product"
    cover["B2"] = "=Inputs!B3"

    inputs = wb.create_sheet("Inputs")
    inputs["A1"] = "Price"
    _paint_input(inputs["B1"], 10)
    inputs["C1"] = "$/sh"
    inputs["A2"] = "Qty"
    _paint_input(inputs["B2"], 3)
    inputs["A3"] = "Product"
    inputs["B3"] = "=B1*B2"
    wb.save(path)


class XlsxModelTests(unittest.TestCase):
    def setUp(self) -> None:
        self.tmp = Path(tempfile.mkdtemp(prefix="xlsx_model_test_"))
        self.xlsx = self.tmp / "Mini Model.xlsx"
        write_mini_workbook(self.xlsx)

    def tearDown(self) -> None:
        shutil.rmtree(self.tmp, ignore_errors=True)

    def test_dump_apply_roundtrip(self) -> None:
        dest = xm.dump_workbook(self.xlsx)
        self.assertTrue(dest.exists())
        self.assertEqual(dest.name, "Mini Model Inputs.md")
        cells = xm.parse_input_tables(dest.read_text(encoding="utf-8"))
        self.assertEqual(len(cells), 2)
        by_coord = {c.coordinate: c for c in cells}
        self.assertEqual(by_coord["B1"].value, 10)
        self.assertEqual(by_coord["B2"].value, 3)

        text = dest.read_text(encoding="utf-8")
        text = text.replace(
            "| Inputs | B1 | Price | 10 | $/sh |",
            "| Inputs | B1 | Price | 20 | $/sh |",
        )
        dest.write_text(text, encoding="utf-8")
        changed, summary = xm.apply_sidecar(self.xlsx)
        self.assertEqual(changed, 1)
        self.assertEqual(summary.get("status"), "success")

        outputs = dict(xm.live_outputs(self.xlsx))
        self.assertEqual(outputs["Price"], 20)
        self.assertEqual(outputs["Product"], 60)

    def test_set_recalculates(self) -> None:
        xm.apply_cells(
            self.xlsx,
            [xm.InputCell(sheet="Inputs", coordinate="B2", value=4)],
        )
        xm.recalc_file(self.xlsx)
        outputs = dict(xm.live_outputs(self.xlsx))
        self.assertEqual(outputs["Product"], 40)

    def test_parse_value_types(self) -> None:
        self.assertEqual(xm.parse_value("31.75"), 31.75)
        self.assertEqual(xm.parse_value("55"), 55)
        self.assertEqual(xm.parse_value("=B6*B7"), "=B6*B7")
        self.assertEqual(xm.parse_value("-0.08"), -0.08)

    def test_float_roundtrip(self) -> None:
        original = 2059.273688604426
        parsed = xm.parse_value(xm.fmt_value(original))
        self.assertTrue(xm.values_close(original, parsed))

    def test_geo_complete_cover_if_present(self) -> None:
        geo = VAULT_GEO
        if not geo.exists():
            self.skipTest("GEO Complete Investment Model.xlsx not in vault")
        outputs = dict(xm.live_outputs(geo))
        self.assertIn("Price", outputs)
        self.assertAlmostEqual(float(outputs["Price"]), 31.75, places=4)
        self.assertAlmostEqual(
            float(outputs["SOTP B (full ICE) $/sh"]), 61.7, delta=0.5
        )


VAULT_GEO = xm.VAULT_ROOT / "Stocks/Infrastructure/GEO/GEO Complete Investment Model.xlsx"


if __name__ == "__main__":
    unittest.main()
